from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from ..extensions import db
from ..models import Cliente, Producto, Venta, Pago, TurnoSesion


ventas_bp = Blueprint('ventas', __name__)

MEDIOS_PAGO_VALIDOS = {'efectivo', 'transferencia', 'tarjeta'}


@ventas_bp.route('/')
@login_required
def listar():
    cliente_id  = request.args.get('cliente_id', type=int)
    medio_pago  = request.args.get('medio_pago', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')

    query = Venta.query.join(Cliente).filter(Cliente.empresa_id == current_user.empresa_id)

    if cliente_id:
        query = query.filter(Venta.cliente_id == cliente_id)
    if fecha_desde:
        try:
            query = query.filter(Venta.fecha >= datetime.fromisoformat(fecha_desde))
        except ValueError:
            flash('Fecha desde inválida.', 'danger')
            return redirect(url_for('ventas.listar'))
    if fecha_hasta:
        try:
            query = query.filter(Venta.fecha <= datetime.fromisoformat(fecha_hasta + 'T23:59:59'))
        except ValueError:
            flash('Fecha hasta inválida.', 'danger')
            return redirect(url_for('ventas.listar'))
    if medio_pago:
        query = query.join(Pago).filter(Pago.medio_pago == medio_pago)

    ventas = query.order_by(Venta.fecha.desc()).all()
    clientes = Cliente.query.filter_by(empresa_id=current_user.empresa_id)\
        .order_by(Cliente.apellido, Cliente.nombre).all()
    total_filtrado = sum(v.total for v in ventas)

    return render_template(
        'ventas/listar.html',
        ventas=ventas,
        clientes=clientes,
        cliente_id=cliente_id,
        medio_pago=medio_pago,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        total_filtrado=total_filtrado,
    )


@ventas_bp.route('/nueva', methods=['GET', 'POST'])
@login_required
def nueva():
    cliente_preseleccionado = request.args.get('cliente_id', type=int)
    clientes = Cliente.query.filter_by(empresa_id=current_user.empresa_id)\
        .order_by(Cliente.apellido, Cliente.nombre).all()
    productos = Producto.query.filter_by(activo=True, empresa_id=current_user.empresa_id)\
        .order_by(Producto.nombre).all()

    if request.method == 'POST':
        medio_pago = request.form['medio_pago']

        try:
            cliente_id  = int(request.form['cliente_id'])
            producto_id = int(request.form['producto_id'])
            monto       = float(request.form['monto'])
        except ValueError:
            flash('Datos inválidos en el formulario.', 'danger')
            return redirect(url_for('ventas.nueva'))

        if medio_pago not in MEDIOS_PAGO_VALIDOS:
            flash('Medio de pago no válido.', 'danger')
            return redirect(url_for('ventas.nueva'))

        if monto < 0:
            flash('El monto no puede ser negativo.', 'danger')
            return redirect(url_for('ventas.nueva'))

        producto = Producto.query.filter_by(id=producto_id, empresa_id=current_user.empresa_id).first_or_404()
        cliente  = Cliente.query.filter_by(id=cliente_id, empresa_id=current_user.empresa_id).first_or_404()

        if monto > producto.precio:
            flash(f'El monto no puede superar el total de la venta (${producto.precio:,.0f}).', 'danger')
            return redirect(url_for('ventas.nueva'))

        cliente.saldo_sesiones += producto.cantidad_sesiones

        venta = Venta(
            cliente_id=cliente_id,
            producto_id=producto_id,
            sesiones_compradas=producto.cantidad_sesiones,
            total=producto.precio,
            fecha=datetime.now(),
        )
        db.session.add(venta)
        db.session.flush()

        pago = Pago(
            venta_id=venta.id,
            medio_pago=medio_pago,
            monto=monto,
            fecha=venta.fecha,
            comprobante=request.form.get('comprobante', '').strip() or None,
        )
        db.session.add(pago)
        db.session.commit()
        flash('Venta y pago registrados correctamente. Sesiones sumadas al cliente.', 'success')
        return redirect(url_for('ventas.listar'))

    return render_template(
        'ventas/form.html',
        clientes=clientes,
        productos=productos,
        cliente_preseleccionado=cliente_preseleccionado,
    )

@ventas_bp.route('/pago_deuda/<int:cliente_id>', methods=['POST'])
@login_required
def pago_deuda(cliente_id):
    cliente    = Cliente.query.filter_by(id=cliente_id, empresa_id=current_user.empresa_id).first_or_404()
    medio_pago = request.form['medio_pago']

    try:
        monto_a_pagar = float(request.form['monto'])
    except ValueError:
        flash('El monto ingresado no es válido.', 'danger')
        return redirect(url_for('clientes.detalle', cliente_id=cliente.id))

    if medio_pago not in MEDIOS_PAGO_VALIDOS:
        flash('Medio de pago no válido.', 'danger')
        return redirect(url_for('clientes.detalle', cliente_id=cliente.id))

    deuda_total = abs(cliente.saldo_pesos)
    if cliente.saldo_pesos >= 0 or monto_a_pagar <= 0:
        flash('No hay deuda para cobrar o el monto es inválido.', 'warning')
        return redirect(url_for('clientes.detalle', cliente_id=cliente.id))

    if monto_a_pagar > deuda_total:
        flash(f'El monto ingresado es mayor a la deuda total (${deuda_total:.2f}).', 'danger')
        return redirect(url_for('clientes.detalle', cliente_id=cliente.id))

    ventas = sorted(cliente.ventas, key=lambda v: v.fecha)
    monto_restante = monto_a_pagar

    for venta in ventas:
        if monto_restante <= 0:
            break
        pagado_en_esta_venta = sum(p.monto for p in venta.pagos)
        falta_pagar_aca = venta.total - pagado_en_esta_venta
        if falta_pagar_aca > 0:
            pago_a_aplicar = min(falta_pagar_aca, monto_restante)
            nuevo_pago = Pago(
                venta_id=venta.id,
                medio_pago=medio_pago,
                monto=pago_a_aplicar,
                comprobante=request.form.get('comprobante', '').strip() or None,
                fecha=datetime.now()
            )
            db.session.add(nuevo_pago)
            monto_restante -= pago_a_aplicar

    db.session.commit()
    flash(f'Se registró un pago de ${monto_a_pagar:.2f} correctamente.', 'success')
    return redirect(url_for('clientes.detalle', cliente_id=cliente.id))


# ─── EXPORTACIÓN A EXCEL ──────────────────────────────────────────────────────

@ventas_bp.route('/exportar/excel')
@login_required
def exportar_excel():
    eid = current_user.empresa_id
    wb  = openpyxl.Workbook()

    estilo_header    = Font(bold=True, color='FFFFFF')
    fill_header      = PatternFill(fill_type='solid', fgColor='2563EB')
    alineacion_centro = Alignment(horizontal='center')

    def crear_hoja(wb, nombre, columnas, filas):
        ws = wb.create_sheet(title=nombre)
        for col_idx, col_nombre in enumerate(columnas, 1):
            celda = ws.cell(row=1, column=col_idx, value=col_nombre)
            celda.font      = estilo_header
            celda.fill      = fill_header
            celda.alignment = alineacion_centro
        for fila_idx, fila in enumerate(filas, 2):
            for col_idx, valor in enumerate(fila, 1):
                ws.cell(row=fila_idx, column=col_idx, value=valor)
        return ws

    clientes = Cliente.query.filter_by(empresa_id=eid).order_by(Cliente.apellido).all()
    crear_hoja(wb, 'Clientes',
        ['ID', 'Apellido', 'Nombre', 'DNI', 'Correo', 'Teléfono', 'Sesiones Restantes', 'Deuda $'],
        [(c.id, c.apellido, c.nombre, c.dni, c.correo, c.telefono,
          c.saldo_sesiones, round(abs(c.saldo_pesos), 2) if c.saldo_pesos < 0 else 0)
         for c in clientes]
    )

    ventas = Venta.query.join(Cliente).filter(Cliente.empresa_id == eid)\
        .order_by(Venta.fecha.desc()).all()
    crear_hoja(wb, 'Ventas',
        ['ID', 'Fecha', 'Cliente', 'Producto', 'Sesiones', 'Total $'],
        [(v.id, v.fecha.strftime('%d/%m/%Y %H:%M'), v.cliente.nombre_completo,
          v.producto.nombre, v.sesiones_compradas, v.total)
         for v in ventas]
    )

    pagos = Pago.query.join(Venta).join(Cliente).filter(Cliente.empresa_id == eid)\
        .order_by(Pago.fecha.desc()).all()
    crear_hoja(wb, 'Pagos',
        ['ID', 'Fecha', 'Cliente', 'Venta ID', 'Medio de Pago', 'Monto $', 'Comprobante'],
        [(p.id, p.fecha.strftime('%d/%m/%Y %H:%M'), p.venta.cliente.nombre_completo,
          p.venta_id, p.medio_pago, p.monto, p.comprobante or '')
         for p in pagos]
    )

    turnos = TurnoSesion.query.join(Cliente).filter(Cliente.empresa_id == eid)\
        .order_by(TurnoSesion.fecha_hora_turno.desc()).all()
    crear_hoja(wb, 'Turnos',
        ['ID', 'Fecha y Hora', 'Cliente', 'Estado', 'Observación'],
        [(t.id, t.fecha_hora_turno.strftime('%d/%m/%Y %H:%M'), t.cliente.nombre_completo,
          t.estado, t.observacion or '')
         for t in turnos]
    )

    productos = Producto.query.filter_by(empresa_id=eid).order_by(Producto.nombre).all()
    crear_hoja(wb, 'Productos',
        ['ID', 'Nombre', 'Descripción', 'Sesiones', 'Precio $', 'Activo'],
        [(p.id, p.nombre, p.descripcion or '', p.cantidad_sesiones, p.precio,
          'Sí' if p.activo else 'No')
         for p in productos]
    )

    turnos_por_cliente = TurnoSesion.query.join(Cliente).filter(Cliente.empresa_id == eid)\
        .order_by(Cliente.apellido, Cliente.nombre, TurnoSesion.fecha_hora_turno.desc()).all()
    crear_hoja(wb, 'Turnos por cliente',
        ['Cliente', 'DNI', 'Fecha y Hora', 'Estado', 'Observación'],
        [(t.cliente.nombre_completo, t.cliente.dni or '',
          t.fecha_hora_turno.strftime('%d/%m/%Y %H:%M'),
          t.estado, t.observacion or '')
         for t in turnos_por_cliente]
    )

    filas_medios = []
    for cliente in clientes:
        totales = {}
        for venta in cliente.ventas:
            for pago in venta.pagos:
                medio = pago.medio_pago
                if medio == 'mercado_pago':
                    medio = 'transferencia'
                medio = medio.replace('_', ' ').capitalize()
                totales[medio] = totales.get(medio, 0) + pago.monto
        if totales:
            for medio, total in sorted(totales.items()):
                filas_medios.append((
                    cliente.nombre_completo,
                    cliente.dni or '',
                    medio,
                    round(total, 2)
                ))

    crear_hoja(wb, 'Medios de pago por cliente',
        ['Cliente', 'DNI', 'Medio de Pago', 'Total Pagado $'],
        filas_medios
    )

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f'cabina_solar_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )